import os
import matlab.engine
import pandas as pd
import shutil
from openpyxl import load_workbook










def read_model(model_path, eng):
    """
    Reads the model from the provided file path and converts it to an Excel file if necessary.
    """
    if model_path.endswith('.mat'):
        eng.eval(f"model = readcbModel('{model_path}')", nargout=0)
        print(f"Model read using readcbModel from .mat file: {model_path}")
        base = os.path.splitext(model_path)[0]
        model_path = f"{base}.xlsx"
        eng.eval(f"writeCBModel(model,'xlsx','{model_path}')", nargout=0)
        
        return model_path

    elif model_path.endswith('.xlsx'):
        eng.eval(f"model = xls2model('{model_path}')", nargout=0)
        print(f"Model read using xls2model from .xlsx file: {model_path}")
        
        return model_path
    else:
        print("Unsupported file type. Please provide .mat or .xlsx files.")

def add_v_to_excel_newfile(excel_path, v_values, new_column_name='v_values', suffix='_modified'):
    """
    Adds a new column of v values to an Excel file, saving it as a new file.
    """
    # Generate new file path
    base = os.path.splitext(excel_path)[0]
    ext = os.path.splitext(excel_path)[1]
    new_path = f"{base}{suffix}{ext}"

    # Copy original file
    if os.path.exists(new_path):
        os.remove(new_path)
    shutil.copyfile(excel_path, new_path)

    wb = load_workbook(new_path)

    # Check if the sheet 'Reaction List' exists
    if 'Reaction List' not in wb.sheetnames:
        wb.close()
        os.remove(new_path)
        raise ValueError("No 'Reaction List' sheet found in the file.")
    
    ws = wb['Reaction List']
    
    # Validate row count matches
    data_row_count = ws.max_row - 1  # Excluding header row
    if len(v_values) != data_row_count:
        wb.close()
        os.remove(new_path)
        raise ValueError(f"Data row count mismatch (v values: {len(v_values)}, sheet rows: {data_row_count})")

    # Insert v values into a new column
    insert_col = 1
    while insert_col <= ws.max_column:
        cell = ws.cell(row=1, column=insert_col)
        if cell.value is None:
            break
        insert_col += 1
    else:
        insert_col = ws.max_column + 1  # Insert at the end if no empty cells

    headers = [cell.value for cell in ws[1]]
    new_col_name = new_column_name
    counter = 1
    while new_col_name in headers:
        new_col_name = f"{new_column_name}_{counter}"
        counter += 1

    ws.cell(row=1, column=insert_col, value=new_col_name)

    # Write the v values to the new column
    sci_format = '0.#####E+00'
    for idx, value in enumerate(v_values, start=2):
        if isinstance(value, matlab.double):
            value = value[0]  # Handle MATLAB double arrays
        
        elif isinstance(value, str) and value.startswith('matlab.double'):
            numeric_part = re.findall(r'\d+\.?\d*', value)[0]
            value = float(numeric_part)
        
        try:
            cell_value = float(value)
        except ValueError as e:
            raise ValueError(f"Cannot convert value to float: {value} (Type: {type(value)})") from e
        
        cell = ws.cell(row=idx, column=insert_col, value=cell_value)
        cell.number_format = sci_format

    wb.save(new_path)
    wb.close()

    # Return the dataframe and file path
    df = pd.read_excel(new_path, sheet_name='Reaction List')
    print(f"New file saved to: {new_path}")
    return df, new_path


def add_optknock_results_to_excel(reaction_list, eng, wb, OptKnockSol,new_col_name):
    """
    Adds the results from the OptKnock optimization to the Excel workbook.
    """
    rxnList = eng.eval("OptKnockSol.rxnList", nargout=1)
    fluxes_values = eng.eval("OptKnockSol.fluxes", nargout=1)
    

    print("rxnList:", rxnList)
   

    # Update reaction list by removing optimized reactions
    for rxn in OptKnockSol['rxnList']:
        if rxn in reaction_list:
            reaction_list.remove(rxn)

    ws = wb['Reaction List']
    insert_col = 1
    while insert_col <= ws.max_column:
        cell = ws.cell(row=1, column=insert_col)
        if cell.value is None:
            break
        insert_col += 1
    else:
        insert_col = ws.max_column + 1  # Insert at the end if no empty cells

    ws.cell(row=1, column=insert_col, value=new_col_name)

    # Write flux values to the new column
    sci_format = '0.#####E+00'
    for idx, value in enumerate(fluxes_values, start=2):
        if isinstance(value, matlab.double):
            value = value[0]  # Handle MATLAB double arrays
        
        elif isinstance(value, str) and value.startswith('matlab.double'):
            numeric_part = re.findall(r'\d+\.?\d*', value)[0]
            value = float(numeric_part)
        
        try:
            cell_value = float(value)
        except ValueError as e:
            raise ValueError(f"Cannot convert value to float: {value} (Type: {type(value)})") from e
        
        cell = ws.cell(row=idx, column=insert_col, value=cell_value)
        cell.number_format = sci_format

    return wb  